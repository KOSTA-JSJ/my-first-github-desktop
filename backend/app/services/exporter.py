"""
Excel 내보내기 서비스
openpyxl을 사용하여 결과를 Excel 파일로 생성합니다.
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from io import BytesIO
from typing import List, Dict


def export_to_excel(results: List[Dict]) -> BytesIO:
    """
    예측 결과를 Excel 파일로 내보냅니다.
    
    Args:
        results: 예측 결과 리스트
        
    Returns:
        Excel 파일의 BytesIO 객체
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "품종 분류 결과"
    
    # 헤더 작성
    headers = ["파일명", "예측 품종", "신뢰도"]
    ws.append(headers)
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 데이터 작성
    for result in results:
        filename = result.get("filename", "")
        prediction = result.get("prediction", "")
        confidence = result.get("confidence", 0.0)
        
        ws.append([filename, prediction, confidence])
    
    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    
    # 신뢰도 컬럼에 조건부 서식 적용
    confidence_col = ws['C']
    
    # 높은 신뢰도 (≥0.8): 녹색
    high_confidence_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    # 중간 신뢰도 (0.5-0.8): 노란색
    medium_confidence_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    # 낮은 신뢰도 (<0.5): 빨간색
    low_confidence_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # 데이터 행에 조건부 서식 적용
    for row_idx in range(2, len(results) + 2):
        confidence_cell = ws[f'C{row_idx}']
        confidence_value = confidence_cell.value
        
        if confidence_value is not None:
            if confidence_value >= 0.8:
                confidence_cell.fill = high_confidence_fill
            elif confidence_value >= 0.5:
                confidence_cell.fill = medium_confidence_fill
            else:
                confidence_cell.fill = low_confidence_fill
            
            # 숫자 포맷 (소수점 3자리)
            confidence_cell.number_format = '0.000'
            confidence_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 파일명 컬럼 정렬
    for row_idx in range(2, len(results) + 2):
        ws[f'A{row_idx}'].alignment = Alignment(horizontal="left", vertical="center")
        ws[f'B{row_idx}'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Excel 파일을 메모리에 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
